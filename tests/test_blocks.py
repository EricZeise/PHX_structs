import random
import warnings

import openpyxl
import pytest

from phx_structs.blocks import find_all_rows_in_col, read_assembly_construction_detail, read_repeating_blocks

warnings.filterwarnings("ignore", category=UserWarning)

HEADER = "Description of building assembly"


@pytest.fixture
def assembly_sheet(tmp_path):
    """A synthetic R-Values-shaped sheet: two populated blocks, one blank slot."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "R-Values"

    # Block 1: header at row 7, identity data at row 8 (populated).
    ws["L7"] = "Description of building assembly"
    ws["Q7"] = "Assembly no."
    ws["L8"] = "Externl wall"
    ws["Q8"] = "01ud"

    # Block 2: header at row 28, identity data at row 29 (populated).
    ws["L28"] = "Description of building assembly"
    ws["Q28"] = "Assembly no."
    ws["L29"] = "Roof"
    ws["Q29"] = "02ud"

    # Block 3: header at row 49, but no display_name at row 50 (unused slot).
    ws["L49"] = "Description of building assembly"
    ws["Q49"] = "Assembly no."

    path = tmp_path / "assembly_sheet.xlsx"
    wb.save(path)
    wb_vals = openpyxl.load_workbook(path, data_only=True)
    wb_fmls = openpyxl.load_workbook(path, data_only=False)
    yield (wb_vals["R-Values"], wb_fmls["R-Values"])
    wb_vals.close()
    wb_fmls.close()


def test_find_all_rows_in_col_finds_every_occurrence(assembly_sheet):
    ws_vals, _ = assembly_sheet
    rows = find_all_rows_in_col(ws_vals, "L", "Description of building assembly")
    assert rows == [7, 28, 49]


def test_find_all_rows_in_col_no_match_returns_empty(assembly_sheet):
    ws_vals, _ = assembly_sheet
    assert find_all_rows_in_col(ws_vals, "L", "Nonexistent header") == []


def test_read_repeating_blocks_skips_unused_slot(assembly_sheet):
    sec_spec = {
        "header_locator": {"col": "L", "string": "Description of building assembly"},
        "items": {"phpp_id_num_col_offset": 5},
        "column_fields": {"display_name": {"column": "L", "unit": None}},
    }
    blocks = read_repeating_blocks(assembly_sheet, sec_spec)
    assert blocks == [
        {"_row": 8, "display_name": "Externl wall", "id": "01ud"},
        {"_row": 29, "display_name": "Roof", "id": "02ud"},
    ]


def test_handles_many_blocks_at_irregular_strides(tmp_path):
    """Real PHPP workbooks vary widely in how many assemblies are populated
    and how they're spaced -- the real sample used a fixed 21-row stride,
    but nothing in find_all_rows_in_col/read_repeating_blocks assumes a fixed
    stride, so this checks a much larger, deliberately irregular layout
    (15-40 rows between blocks, ~1/3 left as unused/blank template slots).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "R-Values"

    rng = random.Random(0)
    row = 7
    expected: list[dict] = []
    n_blocks = 40
    for i in range(n_blocks):
        ws.cell(row=row, column=12, value=HEADER)
        ws.cell(row=row, column=17, value="Assembly no.")
        if i % 3 != 0:  # every 3rd slot left blank/unused, like a real template
            name, ident = f"Assembly_{i:03d}", f"{i:02d}xx"
            ws.cell(row=row + 1, column=12, value=name)
            ws.cell(row=row + 1, column=17, value=ident)
            expected.append({"_row": row + 1, "display_name": name, "id": ident})
        row += rng.randint(15, 40)

    path = tmp_path / "many_blocks.xlsx"
    wb.save(path)
    wb_vals = openpyxl.load_workbook(path, data_only=True)
    wb_fmls = openpyxl.load_workbook(path, data_only=False)
    try:
        ws_pair = (wb_vals["R-Values"], wb_fmls["R-Values"])
        header_rows = find_all_rows_in_col(wb_vals["R-Values"], "L", HEADER)
        assert len(header_rows) == n_blocks

        sec_spec = {
            "header_locator": {"col": "L", "string": HEADER},
            "items": {"phpp_id_num_col_offset": 5},
            "column_fields": {"display_name": {"column": "L", "unit": None}},
        }
        assert read_repeating_blocks(ws_pair, sec_spec) == expected
    finally:
        wb_vals.close()
        wb_fmls.close()


def test_no_header_occurrences_returns_empty_list(tmp_path):
    """A workbook/sheet with no matching header at all (e.g. wrong sheet,
    or a PHPP version with different wording) must not error -- just
    produce zero blocks."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "R-Values"
    ws["L7"] = "Some unrelated label"

    path = tmp_path / "no_header.xlsx"
    wb.save(path)
    wb_vals = openpyxl.load_workbook(path, data_only=True)
    wb_fmls = openpyxl.load_workbook(path, data_only=False)
    try:
        ws_pair = (wb_vals["R-Values"], wb_fmls["R-Values"])
        sec_spec = {
            "header_locator": {"col": "L", "string": HEADER},
            "items": {"phpp_id_num_col_offset": 5},
            "column_fields": {"display_name": {"column": "L", "unit": None}},
        }
        assert read_repeating_blocks(ws_pair, sec_spec) == []
    finally:
        wb_vals.close()
        wb_fmls.close()


# --- read_assembly_construction_detail: full R-value/layer resolution ---

_CONSTRUCTOR_SEC_SPEC = {
    "header_locator": {"col": "L", "string": HEADER},
    "items": {
        "phpp_id_num_col_offset": 5,
        "name_row_offset": 2,
        "rsi_row_offset": 4,
        "rse_row_offset": 5,
        "first_layer_row_offset": 7,
        "last_layer_row_offset": 14,
        "result_val_row_offset": 19,
        "result_val_col": "R",
    },
    "column_fields": {
        "display_name": {"column": "L"},
        "r_si": {"column": "M"},
        "r_se": {"column": "M"},
        "interior_insulation": {"column": "R"},
        "sec_1_description": {"column": "L"},
        "sec_1_conductivity": {"column": "M"},
        "sec_2_description": {"column": "N"},
        "sec_2_conductivity": {"column": "O"},
        "sec_3_description": {"column": "P"},
        "sec_3_conductivity": {"column": "Q"},
        "thickness": {"column": "R"},
        "u_val_supplement": {"column": "R"},
        "sec_2_percentage": {"column": "O"},
        "sec_3_percentage": {"column": "Q"},
    },
}


@pytest.fixture
def construction_detail_sheet(tmp_path):
    """A single assembly block laid out exactly like the real "Externl wall"
    block confirmed against Example_IP.xlsx: header at row 7 (offset 0),
    display_name/id at row 8 (offset 1), rsi/rse dropdown cells at rows
    10/11 (offsets 3/4), a 3-layer table at rows 13-15 (offsets 6-8, one
    layer with a sec_2 sub-material, one trailing blank slot at offset 9
    to confirm the table stops there), percentages at row 21 (offset 14),
    and the R-value result at row 25 (offset 18).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "R-Values"

    ws["L7"], ws["Q7"] = HEADER, "Assembly no."
    ws["L8"], ws["Q8"] = "Test Wall", "01ud"
    ws["L10"], ws["M10"] = "Orientation of building assembly (or Rsi)", "2-Wall"
    ws["L11"], ws["M11"] = "Adjacent to (or Rsa)", "1-Outdoor air"

    ws["L13"], ws["M13"], ws["R13"] = "Gypsum", 0.9, 0.625
    ws["L14"], ws["M14"], ws["R14"] = "Cellulose", 3.6, 3.5
    ws["N14"], ws["O14"] = "2x4 @ 24\"", 1.3  # sec_2 sub-material on this layer only
    ws["L15"], ws["M15"], ws["R15"] = "Sheathing", 1.4, 0.625
    # row 16 (offset 9) intentionally blank -- the layer table must stop here,
    # not continue reading through the remaining offset-10-through-13 slots.

    ws["L21"], ws["O21"] = "Percentage of sec. 1:", 0.094

    ws["Q25"], ws["R25"] = "R-value [...]:", 41.64

    path = tmp_path / "construction_detail.xlsx"
    wb.save(path)
    wb_vals = openpyxl.load_workbook(path, data_only=True)
    wb_fmls = openpyxl.load_workbook(path, data_only=False)
    yield (wb_vals["R-Values"], wb_fmls["R-Values"])
    wb_vals.close()
    wb_fmls.close()


def test_construction_detail_resolves_scalar_fields(construction_detail_sheet):
    blocks = read_assembly_construction_detail(construction_detail_sheet, _CONSTRUCTOR_SEC_SPEC)
    assert len(blocks) == 1
    block = blocks[0]
    assert block["display_name"] == "Test Wall"
    assert block["id"] == "01ud"
    assert block["r_si"] == "2-Wall"
    assert block["r_se"] == "1-Outdoor air"
    assert block["result_val"] == 41.64


def test_construction_detail_resolves_layers_and_stops_at_first_blank(construction_detail_sheet):
    blocks = read_assembly_construction_detail(construction_detail_sheet, _CONSTRUCTOR_SEC_SPEC)
    layers = blocks[0]["layers"]
    assert len(layers) == 3  # not 8 -- must stop at the first blank slot, not read the full window
    assert layers[0]["sec_1_description"] == "Gypsum"
    assert layers[0]["sec_1_conductivity"] == 0.9
    assert layers[0]["thickness"] == 0.625
    assert "sec_2_description" not in layers[0]


def test_construction_detail_includes_sec_2_only_when_present(construction_detail_sheet):
    blocks = read_assembly_construction_detail(construction_detail_sheet, _CONSTRUCTOR_SEC_SPEC)
    layers = blocks[0]["layers"]
    assert layers[1]["sec_1_description"] == "Cellulose"
    assert layers[1]["sec_2_description"] == "2x4 @ 24\""
    assert layers[1]["sec_2_conductivity"] == 1.3
    assert "sec_3_description" not in layers[1]
    # third layer (Sheathing) has neither sec_2 nor sec_3
    assert "sec_2_description" not in layers[2]
    assert "sec_3_description" not in layers[2]


def test_construction_detail_post_layer_percentage(construction_detail_sheet):
    blocks = read_assembly_construction_detail(construction_detail_sheet, _CONSTRUCTOR_SEC_SPEC)
    assert blocks[0]["sec_2_percentage"] == 0.094
    assert blocks[0]["sec_3_percentage"] is None


def test_construction_detail_unset_scalars_are_none(construction_detail_sheet):
    blocks = read_assembly_construction_detail(construction_detail_sheet, _CONSTRUCTOR_SEC_SPEC)
    assert blocks[0]["interior_insulation"] is None
    assert blocks[0]["u_val_supplement"] is None


def test_construction_detail_skips_unused_slot(assembly_sheet):
    """assembly_sheet's third block (row 49) has no display_name -- must
    still be skipped here exactly like read_repeating_blocks, and the two
    real blocks resolved even though this fixture has no layer/rsi/rse data
    at all (every construction-detail field should just come back None)."""
    blocks = read_assembly_construction_detail(assembly_sheet, _CONSTRUCTOR_SEC_SPEC)
    assert len(blocks) == 2
    assert blocks[0]["display_name"] == "Externl wall"
    assert blocks[0]["layers"] == []
    assert blocks[0]["r_si"] is None
