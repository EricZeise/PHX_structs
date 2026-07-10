"""Joins Components/Windows/Areas/HVAC/thermal-bridges by PHPP's reference conventions.

WINDOWS.window_rows[].frame_id/glazing_id, AREAS.surface_rows[].assembly_id,
and ADDNL_VENT.units[].unit_selected are composite strings PHPP itself builds
for its dropdowns (e.g. "01ud-Climatop Ultra N") -- not bare IDs. Splitting on
the first "-" and matching the prefix against the relevant COMPONENTS/assembly
library resolved 128/128 real window references with zero mismatches during
this tool's feasibility check, and the same mechanism resolves ventilator
selections.

ADDNL_VENT uses a second, different reference convention for its own
rooms<->units<->ducts chain: not an id string, but a 1-based *ordinal
position* in a sibling list -- ADDNL_VENT.rooms[].vent_unit_assigned is an
int naming which row of ADDNL_VENT.units it uses (confirmed empirically:
5 rooms, 5 units, room i+1 <-> units[i], exact 1:1), and
ADDNL_VENT.ducts[].duct_assign_1..duct_assign_10 are ten flag columns where
a truthy value in column N means the duct serves ADDNL_VENT.rooms[N-1].
resolve_ordinal() is the primitive for this second convention.

AREAS.thermal_bridge_rows carries no id-style reference to any other
worksheet at all -- it's standalone data. Its group_number field happens to
share the same "<code>-<label>" text shape as surface_rows.group_number, but
the two use disjoint code families (thermal bridges: e.g. "17-...";
surfaces: e.g. "10-...", "11-...") confirmed against a real workbook, so
this is a category label to group by, not a cross-reference to resolve.

Any reference that fails to resolve is recorded under "unresolved" rather
than silently dropped -- the same philosophy phpp-shape-sync's normalizer.py
uses for locators it can't resolve against a real workbook.
"""

from __future__ import annotations

import warnings
from typing import Any

from phx_structs.blocks import read_assembly_construction_detail
from phx_structs.sibling_import import ensure_phpp_tool_importable

ensure_phpp_tool_importable()

from phpp_tool.locators import resolve_sheet_name  # noqa: E402
from phpp_tool.map_parser import parse_field_map  # noqa: E402
from phpp_tool.reader import read_phpp  # noqa: E402


def _split_ref(value: Any) -> tuple[str, str] | None:
    """Split a "<id>-<description>" reference into (id, description).

    Returns None for a value that isn't a non-empty string containing "-".
    """
    if not isinstance(value, str) or "-" not in value:
        return None
    ref_id, _, description = value.partition("-")
    return ref_id, description


def _index_by_id(records: list[dict[str, Any]] | None) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for record in records or []:
        rid = record.get("id")
        if rid:
            index[rid] = record
    return index


def resolve_reference(
    row: dict[str, Any], field_name: str, index: dict[str, dict[str, Any]],
) -> str | None:
    """Resolve row[field_name] (an "<id>-<description>" string) against *index*.

    Returns the matched id, or None if the field is absent/unparseable or
    doesn't match any key in *index* -- callers are responsible for routing
    a None result to an "unresolved" list rather than dropping it silently.
    """
    split = _split_ref(row.get(field_name))
    if split is None:
        return None
    ref_id, _ = split
    return ref_id if ref_id in index else None


def resolve_ordinal(position: Any, records: list[dict[str, Any]]) -> dict[str, Any] | None:
    """Resolve a 1-based ordinal position (ADDNL_VENT's rooms<->units<->ducts
    convention) against a sibling list.

    Returns the record at records[position - 1], or None if *position* isn't
    a positive int (or int-valued float, as PHPP scalars often come through
    openpyxl) or falls outside the list -- callers route a None to
    "unresolved" the same as resolve_reference.
    """
    if isinstance(position, float) and position.is_integer():
        position = int(position)
    if not isinstance(position, int) or isinstance(position, bool) or position < 1:
        return None
    idx = position - 1
    if idx >= len(records):
        return None
    return records[idx]


def _read_assemblies(workbook_path: str, field_map_path: str) -> list[dict[str, Any]]:
    import openpyxl

    field_map = parse_field_map(field_map_path)
    uvalues = field_map.get("UVALUES")
    if uvalues is None:
        return []
    constructor = uvalues.get("sections", {}).get("constructor")
    if constructor is None:
        return []

    warnings.filterwarnings("ignore", category=UserWarning)
    wb_vals = openpyxl.load_workbook(workbook_path, data_only=True)
    wb_fmls = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        real_name = resolve_sheet_name(uvalues["sheet_name"], wb_vals.sheetnames)
        if real_name is None:
            return []
        ws_pair = (wb_vals[real_name], wb_fmls[real_name])
        return read_assembly_construction_detail(ws_pair, constructor, skip_formulas=False)
    finally:
        wb_vals.close()
        wb_fmls.close()


def _build_hvac(data: dict[str, Any], ventilators_by_id: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], dict[str, list]]:
    """Join ADDNL_VENT's rooms<->units<->ducts chain (see module docstring).

    units[].unit_selected resolves against ventilators_by_id via the
    id-string convention (resolve_reference); rooms[].vent_unit_assigned and
    ducts[].duct_assign_N resolve via 1-based ordinal position
    (resolve_ordinal) against the sibling units/rooms lists respectively --
    a duct can be assigned to more than one room, so its result is a list.
    """
    addnl_vent = data.get("ADDNL_VENT") or {}
    units = addnl_vent.get("units") or []
    rooms = addnl_vent.get("rooms") or []
    ducts = addnl_vent.get("ducts") or []

    unresolved_units: list[dict[str, Any]] = []
    unresolved_rooms: list[dict[str, Any]] = []
    unresolved_ducts: list[dict[str, Any]] = []

    units_out: list[dict[str, Any]] = []
    for row in units:
        entry = dict(row)
        if row.get("unit_selected") is not None:
            resolved = resolve_reference(row, "unit_selected", ventilators_by_id)
            if resolved is not None:
                entry["resolved_ventilator_id"] = resolved
            else:
                unresolved_units.append({
                    "_row": row.get("_row"), "field": "unit_selected", "value": row.get("unit_selected"),
                })
        units_out.append(entry)

    rooms_out: list[dict[str, Any]] = []
    for row in rooms:
        entry = dict(row)
        position = row.get("vent_unit_assigned")
        if position is not None:
            unit = resolve_ordinal(position, units)
            if unit is not None:
                entry["resolved_unit_row"] = unit.get("_row")
                entry["resolved_unit_display_name"] = unit.get("display_name")
            else:
                unresolved_rooms.append({
                    "_row": row.get("_row"), "field": "vent_unit_assigned", "value": position,
                })
        rooms_out.append(entry)

    ducts_out: list[dict[str, Any]] = []
    for row in ducts:
        entry = dict(row)
        resolved_rooms: list[dict[str, Any]] = []
        for position in range(1, 11):
            flag = row.get(f"duct_assign_{position}")
            if not flag:
                continue
            room = resolve_ordinal(position, rooms)
            if room is not None:
                resolved_rooms.append({"_row": room.get("_row"), "display_name": room.get("display_name")})
            else:
                unresolved_ducts.append({
                    "_row": row.get("_row"), "field": f"duct_assign_{position}", "value": flag,
                })
        if resolved_rooms:
            entry["resolved_rooms"] = resolved_rooms
        ducts_out.append(entry)

    hvac = {"units": units_out, "rooms": rooms_out, "ducts": ducts_out}
    unresolved = {"hvac_units": unresolved_units, "hvac_rooms": unresolved_rooms, "hvac_ducts": unresolved_ducts}
    return hvac, unresolved


def _build_thermal_bridges(data: dict[str, Any]) -> list[dict[str, Any]]:
    """Parse AREAS.thermal_bridge_rows' group_number into (code, label).

    Not a cross-reference resolution -- see module docstring: thermal
    bridges carry no id-style link to any other worksheet, this only
    structures the existing "<code>-<label>" text for easier grouping.
    """
    rows = (data.get("AREAS") or {}).get("thermal_bridge_rows") or []
    out: list[dict[str, Any]] = []
    for row in rows:
        entry = dict(row)
        split = _split_ref(row.get("group_number"))
        if split is not None:
            entry["group_code"], entry["group_label"] = split
        out.append(entry)
    return out


def list_assemblies(crossref: dict[str, Any]) -> list[dict[str, Any]]:
    """Group every data field associated with each assembly into one record per assembly.

    Takes build_crossref()'s output and, for each resolved assembly, bundles
    its own construction-detail fields (identity, R/U-value, layers -- already
    present in `assemblies`, see blocks.read_assembly_construction_detail)
    together with every AREAS.surface_rows entry that references it (via
    resolved_assembly_id). This is the "list assemblies and group their
    associated data" view -- build_crossref() itself keeps assemblies and
    areas as separate flat structures, cross-referenced only by id.

    Assemblies with no referencing surface rows are still included, with an
    empty "areas" list and area_total 0.0 -- an unused library entry is real
    information (e.g. a template slot nobody assigned), not something to hide.

    Ordering follows dict insertion order, i.e. the order assemblies were
    found on the R-Values sheet (top to bottom), not alphabetical by id.
    """
    assemblies = crossref.get("assemblies", {})
    areas = crossref.get("areas", [])

    areas_by_assembly: dict[str, list[dict[str, Any]]] = {aid: [] for aid in assemblies}
    for area in areas:
        aid = area.get("resolved_assembly_id")
        if aid in areas_by_assembly:
            areas_by_assembly[aid].append(area)

    result: list[dict[str, Any]] = []
    for aid, assembly in assemblies.items():
        used_by = areas_by_assembly[aid]
        record = dict(assembly)
        record["areas"] = used_by
        record["area_total"] = sum(a.get("area") or 0 for a in used_by)
        result.append(record)
    return result


def build_crossref(workbook_path: str, field_map_path: str) -> dict[str, Any]:
    """Read *workbook_path* and return the joined Components/Windows/Areas/HVAC structure."""
    data = read_phpp(workbook_path, field_map_path, skip_formulas=False)

    frames = (data.get("COMPONENTS") or {}).get("frames") or []
    glazings = (data.get("COMPONENTS") or {}).get("glazings") or []
    ventilators = (data.get("COMPONENTS") or {}).get("ventilators") or []
    window_rows = (data.get("WINDOWS") or {}).get("window_rows") or []
    surface_rows = (data.get("AREAS") or {}).get("surface_rows") or []
    assemblies = _read_assemblies(workbook_path, field_map_path)

    frames_by_id = _index_by_id(frames)
    glazings_by_id = _index_by_id(glazings)
    ventilators_by_id = _index_by_id(ventilators)
    assemblies_by_id = _index_by_id(assemblies)

    windows_out: list[dict[str, Any]] = []
    unresolved_windows: list[dict[str, Any]] = []
    for row in window_rows:
        entry = dict(row)
        for field_name, index in (("frame_id", frames_by_id), ("glazing_id", glazings_by_id)):
            if row.get(field_name) is None:
                continue
            resolved = resolve_reference(row, field_name, index)
            if resolved is not None:
                entry[f"resolved_{field_name}"] = resolved
            else:
                unresolved_windows.append({
                    "_row": row.get("_row"), "field": field_name, "value": row.get(field_name),
                })
        windows_out.append(entry)

    areas_out: list[dict[str, Any]] = []
    unresolved_areas: list[dict[str, Any]] = []
    for row in surface_rows:
        entry = dict(row)
        if row.get("assembly_id") is not None:
            resolved = resolve_reference(row, "assembly_id", assemblies_by_id)
            if resolved is not None:
                entry["resolved_assembly_id"] = resolved
            else:
                unresolved_areas.append({
                    "_row": row.get("_row"), "field": "assembly_id", "value": row.get("assembly_id"),
                })
        areas_out.append(entry)

    hvac, hvac_unresolved = _build_hvac(data, ventilators_by_id)
    thermal_bridges = _build_thermal_bridges(data)

    return {
        "components": {"frames": frames_by_id, "glazings": glazings_by_id, "ventilators": ventilators_by_id},
        "assemblies": assemblies_by_id,
        "windows": windows_out,
        "areas": areas_out,
        "thermal_bridges": thermal_bridges,
        "hvac": hvac,
        "unresolved": {
            "windows": unresolved_windows,
            "areas": unresolved_areas,
            **hvac_unresolved,
        },
    }
